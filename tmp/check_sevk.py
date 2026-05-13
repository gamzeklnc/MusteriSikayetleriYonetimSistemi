import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\yasin\MusteriSikayetleriYonetimSistemi\Excel_Aktar\Sevk Edilenler - Güncel.xlsx', data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")

# Count actual rows with data
actual_rows = 0
for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
    if any(cell is not None and str(cell).strip() for cell in row):
        actual_rows += 1
    else:
        break  # Stop at first empty row

print(f"Rows with data (excluding header): {actual_rows}")

# Print headers
print("\n--- Headers ---")
for i, cell in enumerate(ws[1], 1):
    print(f"  Col {i} ({cell.column_letter}): {cell.value}")

# Print sample data rows
print("\n--- Sample rows 2-8 ---")
for row_num in range(2, min(9, actual_rows + 2)):
    musteri = ws.cell(row_num, 3).value
    sevk_tarihi = ws.cell(row_num, 2).value
    sevk_adet = ws.cell(row_num, 7).value
    fabrika = ws.cell(row_num, 11).value
    print(f"Row {row_num}: Musteri={musteri}, SevkTarihi={sevk_tarihi}, SevkAdet={sevk_adet}, Fabrika={fabrika}")

# Get unique customer names
customers = set()
for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[2] and str(row[2]).strip():
        customers.add(str(row[2]).strip())
    else:
        break

print(f"\n--- Unique customers (first batch): {len(customers)} ---")
for c in sorted(list(customers))[:15]:
    print(f"  - {c}")
